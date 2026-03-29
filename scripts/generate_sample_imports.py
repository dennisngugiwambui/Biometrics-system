"""
Generate sample Excel templates for Student and Teacher bulk import.

Creates two .xlsx files with Kenyan names (Kikuyu, Kalenjin, Luo, Luhya, Kamba, etc.)
for schools to use as templates and sample data. Column headers match the backend
bulk import API (CSV/Excel column names are normalized to snake_case).
"""

from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    raise

# Output directory: project root (both Excel sheets saved here)
PROJECT_ROOT = Path(__file__).resolve().parent.parent

# --- Kenyan names by community (first name, last name) ---
# Kikuyu, Kalenjin, Luo, Luhya, Kamba, Kisii, Maasai, Meru, etc.
STUDENTS = [
    # admission_number, first_name, last_name, date_of_birth, gender, parent_phone, parent_email, class_name, stream_name, is_boarding
    ("ADM-2024-001", "Wanjiru", "Mwangi", "2012-03-15", "female", "+254712345001", "wanjiru.parent@email.com", "Grade 6", "East", "false"),
    ("ADM-2024-002", "Kipchoge", "Chebet", "2011-07-22", "male", "+254723456002", "", "Grade 7", "West", "true"),
    ("ADM-2024-003", "Omondi", "Akinyi", "2013-01-08", "male", "+254734567003", "akinyi.g@email.com", "Grade 5", "North", "false"),
    ("ADM-2024-004", "Njeri", "Kamau", "2012-11-30", "female", "+254745678004", "", "Grade 6", "South", "false"),
    ("ADM-2024-005", "Wekesa", "Nasimiyu", "2011-09-14", "male", "+254756789005", "wekesa.parent@email.com", "Grade 7", "East", "true"),
    ("ADM-2024-006", "Muthoki", "Kioko", "2013-05-03", "female", "+254767890006", "", "Grade 5", "West", "false"),
    ("ADM-2024-007", "Kerubo", "Nyambane", "2012-08-19", "female", "+254778901007", "kerubo.parent@email.com", "Grade 6", "North", "false"),
    ("ADM-2024-008", "Otieno", "Odhiambo", "2011-12-25", "male", "+254789012008", "", "Grade 7", "South", "true"),
    ("ADM-2024-009", "Naserian", "Sironka", "2013-02-11", "female", "+254790123009", "naserian.parent@email.com", "Grade 5", "East", "false"),
    ("ADM-2024-010", "Muthoni", "Kinyua", "2012-04-27", "female", "+254701234010", "", "Grade 6", "West", "false"),
]

TEACHERS = [
    # first_name, last_name, phone, email, subject, department
    ("James", "Kariuki", "+254712345101", "j.kariuki@school.ke", "Mathematics", "Sciences"),
    ("Grace", "Jepkosgei", "+254723456102", "g.jepkosgei@school.ke", "English", "Languages"),
    ("Peter", "Ochieng", "+254734567103", "p.ochieng@school.ke", "Kiswahili", "Languages"),
    ("Mary", "Wambui", "+254745678104", "m.wambui@school.ke", "Science", "Sciences"),
    ("David", "Barasa", "+254756789105", "d.barasa@school.ke", "Social Studies", "Humanities"),
]

# Header styling
HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _write_instructions(ws, title: str, columns: list[str], note: str):
    """Write instruction rows at the top (row 1)."""
    ws.merge_cells("A1:" + get_column_letter(len(columns)) + "1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A2:" + get_column_letter(len(columns)) + "2")
    ws["A2"] = note
    ws["A2"].font = Font(italic=True, size=9)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18


def _style_header_row(ws, row: int, columns: list[str]):
    for c, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=c, value=col_name)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _write_upload_ready_sheet(wb, sheet_name: str, columns: list[str], rows: list):
    """Sheet with only header + data for saving as CSV and uploading."""
    ws = wb.create_sheet(sheet_name)
    for c, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT_WHITE
        cell.border = THIN_BORDER
    for r, row_data in enumerate(rows, start=2):
        for c, value in enumerate(row_data, start=1):
            ws.cell(row=r, column=c, value=value).border = THIN_BORDER
    for c in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(c)].width = max(14, len(columns[c - 1]) + 2)
    return ws


def generate_students_template():
    """Generate students bulk import sample Excel."""
    columns = [
        "admission_number",
        "first_name",
        "last_name",
        "date_of_birth",
        "gender",
        "parent_phone",
        "parent_email",
        "class_name",
        "stream_name",
        "is_boarding",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    _write_instructions(
        ws,
        "Student bulk import template",
        columns,
        "Use this sheet as template. Headers must match exactly. date_of_birth: YYYY-MM-DD. gender: male/female/other. is_boarding: true/false. class_name and stream_name must exist in your school.",
    )
    header_row = 3
    _style_header_row(ws, header_row, columns)
    for r, row_data in enumerate(STUDENTS, start=header_row + 1):
        for c, value in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.border = THIN_BORDER
            if c == 4:
                cell.number_format = "YYYY-MM-DD"
    for c in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(c)].width = max(14, len(columns[c - 1]) + 2)

    _write_upload_ready_sheet(wb, "Upload_Ready", columns, STUDENTS)

    out_path = PROJECT_ROOT / "students_import_template.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")
    return out_path


def generate_teachers_template():
    """Generate teachers bulk import sample Excel."""
    columns = [
        "first_name",
        "last_name",
        "phone",
        "email",
        "subject",
        "department",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Teachers"
    _write_instructions(
        ws,
        "Teacher bulk import template",
        columns,
        "Use this sheet as template. Headers must match exactly. phone must be unique per school. Save the 'Upload_Ready' sheet as CSV to upload in the app.",
    )
    header_row = 3
    _style_header_row(ws, header_row, columns)
    for r, row_data in enumerate(TEACHERS, start=header_row + 1):
        for c, value in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.border = THIN_BORDER
    for c in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(c)].width = max(14, len(columns[c - 1]) + 2)

    _write_upload_ready_sheet(wb, "Upload_Ready", columns, TEACHERS)

    out_path = PROJECT_ROOT / "teachers_import_template.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")
    return out_path


def main():
    generate_students_template()
    generate_teachers_template()
    print("Done. Files are in project root:", PROJECT_ROOT)


if __name__ == "__main__":
    main()
