"""
Parse uploaded bulk-import files (CSV, TSV, or XLSX) into a list of dicts with normalized keys.

Used by students and teachers import/file endpoints. Supports .csv, .tsv, .txt, and .xlsx.
"""

import csv
import io
from typing import List, Dict, Any

# Normalize key for lookup: strip, lower, replace spaces with underscore
def _norm_key(k: str) -> str:
    return (k or "").strip().lower().replace(" ", "_")


def _normalize_row(raw: Dict[str, Any]) -> Dict[str, str]:
    return {_norm_key(k): (v if v is None else str(v).strip()) for k, v in raw.items()}


def _parse_csv_tsv(content: bytes, filename: str) -> List[Dict[str, str]]:
    try:
        text = content.decode("utf-8")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    # Normalize line endings to avoid "new-line character seen in unquoted field" (Excel \r\n)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except Exception:
        dialect = "excel"
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    out: List[Dict[str, str]] = []
    for raw in reader:
        out.append(_normalize_row(raw))
    return out


def _parse_xlsx(content: bytes, filename: str) -> List[Dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError("XLSX support requires openpyxl. Install with: pip install openpyxl")
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    # Prefer Upload_Ready, then sheet named Teachers/Students, then first sheet
    sheet = None
    for name in ("Upload_Ready", "Teachers", "Students"):
        if name in wb.sheetnames:
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    raw_headers = rows[0]
    headers = []
    for i, h in enumerate(raw_headers):
        key = _norm_key(str(h or ""))
        if key:
            headers.append((i, key))
    out: List[Dict[str, str]] = []
    for row in rows[1:]:
        d = {
            key: (str(row[i]).strip() if i < len(row) and row[i] is not None else "")
            for i, key in headers
        }
        if any(d.values()):  # skip completely empty rows
            out.append(d)
    return out


def parse_upload_to_rows(content: bytes, filename: str) -> List[Dict[str, str]]:
    """
    Parse uploaded file (CSV, TSV, or XLSX) into list of dicts with normalized keys.

    :param content: Raw file bytes
    :param filename: Original filename (used to choose parser)
    :return: List of dicts; keys are normalized (lowercase, spaces -> underscores)
    :raises ValueError: If format is unsupported or parsing fails
    """
    name = (filename or "").lower().strip()
    if name.endswith(".xlsx") or name.endswith(".xls"):
        return _parse_xlsx(content, filename)
    if name.endswith(".csv") or name.endswith(".tsv") or name.endswith(".txt"):
        return _parse_csv_tsv(content, filename)
    # By content: XLSX starts with PK (zip)
    if content[:2] == b"PK":
        return _parse_xlsx(content, filename)
    # Default: try as CSV/TSV (e.g. no extension)
    return _parse_csv_tsv(content, filename)
